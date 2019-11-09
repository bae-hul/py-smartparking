import serial #Serial imported for Serial communication
import time #Required to use delay functions
from twilio.rest import Client
import datetime
import openpyxl


class project(object):
    def freerowcheck(self,col):
            i=1
            temp=" "
            while temp!=None:
                i=i+1
                temp=self.sheet_obj[col+str(i)].value
            return i

    def sendsms(self):
        parking=999
        for i in range(2):
            if self.ps[i] == 0:
                
                parking=i+1
                
                
        self.ps[i]=2
        self.sheet_obj["P"+str(parking)].value=2
        self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
        account_sid = 'YOUR USERNAME'
        auth_token = 'YOUR AUTH TOKEN'
        client = Client(account_sid, auth_token)
        """
        if parking!=999:
            message = client.messages \
                            .create(
                                 body=("Hello "+str(self.name)+"! A free parking has been assinged to you: " + str(parking)),
                                 from_='XXX',
                                 to='XXX'
                             )
            print("SMS sent to user regarding parking"+str(parking))
            
        else:
            message = client.messages \
                            .create(
                                 body=("Sorry. Parking is full."),
                                 from_='XXX',
                                 to='XXX'
                             )
            print("Parking full!")
         """
        print("SMS")
        
        
        
        self.ArduinoSerial.reset_input_buffer()
        self.ArduinoSerial.flush()
        self.ArduinoSerial.reset_output_buffer()
        self.ArduinoSerial.readline()
        self.ArduinoSerial.close()
        self.mainmenu()
        
        
    def rfid(self,temp):
        #b'ID:  AB 13 06 7512\r\n'
        temp=temp[7:18]
        temp=temp.replace(" ","")
        print(temp)
        self.name=self.rowc(temp)
        
        if self.rowval!=0:
            self.entry()
        #ArduinoSerial.write("USD200".encode())
        else:
            self.ArduinoSerial.write("Hello Guest".encode())
            time.sleep(1)
            self.ArduinoSerial.write("Plz Register".encode())
            
            
        time.sleep(5)

    def entry(self):
        datetimeFormat = '%Y-%m-%d %H:%M:%S.%f'
        if self.sheet_obj["F"+str(self.rowval)].value==0 or self.sheet_obj["F"+str(self.rowval)].value==None:
            self.sheet_obj["D"+str(self.rowval)].value=datetime.datetime.now()
            self.sheet_obj["F"+str(self.rowval)].value=1
            name="Hello "+self.name
            print(name)
            self.ArduinoSerial.write(name.encode())
            time.sleep(1)
            credit="Credit: Rs"+str(self.sheet_obj["I"+str(self.rowval)].value)
            self.ArduinoSerial.write(credit.encode())
            self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
            self.sendsms()
        
            
        elif self.sheet_obj["F"+str(self.rowval)].value==1:
            self.sheet_obj["E"+str(self.rowval)].value=datetime.datetime.now()
            s=datetime.datetime.now()
            diff = datetime.datetime.strptime(str(s), datetimeFormat)- datetime.datetime.strptime(str(self.sheet_obj["D"+str(self.rowval)].value), datetimeFormat)
            name="Bye "+self.name
            self.ArduinoSerial.write(name.encode())
            time.sleep(1)
            print("Difference:", int(diff.seconds)//60)
            self.sheet_obj["G"+str(self.rowval)].value=int(diff.seconds)/60
            self.sheet_obj["F"+str(self.rowval)].value=0
            
            time1="Parked: "+str(int(diff.seconds)//60)+"Mins"
            self.ArduinoSerial.write(time1.encode())
        self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
            
        

    def rowc(self,uid):
        i=1
        temp=" "
        while temp!=None:
            i=i+1
            temp=self.sheet_obj["A"+str(i)].value
            if temp==uid:
                self.rowval=i
                return self.sheet_obj["B"+str(i)].value
        return "Guest"
        
        
    def __init__ (self):
        self.mode=0
        self.acc=" "
        self.auth=[]
        self.username=" "
        self.otp=0
        self.cost=0
        self.wb=" "
        self.sheet_obj=" "
        self.wb_obj=" "
        self.location=" "
        self.access=" "
        self.rowval=0
        self.name=""
        self.ps=[]
        
    
#Main

    def mainmenu(self):
        ArduinoSerial = serial.Serial('com11',9600) #Create Serial port object called arduinoSerialData
        self.ArduinoSerial = ArduinoSerial
        time.sleep(2) #wait for 2 secounds for the communication to get established

        try:
            path="C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx"
            self.wb_obj = openpyxl.load_workbook(path)
            self.sheet_obj = self.wb_obj.active
            self.ps=[self.sheet_obj["P1"].value]
            self.ps.append(self.sheet_obj["P2"].value)
            print("Success")
        except:
            wb = openpyxl.Workbook()
            sheet=wb.active
            sheet['A1'].value="UID"
            sheet['B1'].value="Name"
            sheet['D1'].value="Entry"
            sheet['E1'].value="Exit"
            sheet['G1'].value="Diff"
            sheet['F1'].value="Status"
            sheet['I1'].value="Credit"
            wb.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
            path="C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx"
            self.wb_obj = openpyxl.load_workbook(path)
            self.sheet_obj = self.wb_obj.active
            

        #print ("Test program - Arduino to Python connection.")
        p1c=100
        p2c=100
        flag=-1
        p1flag=0
        p2flag=0
        totalfree=0
        flag2=0
        while 1: #Do this forever
            print(self.ps)
            status=str(totalfree)
            #ArduinoSerial.write(status.encode())
            if flag<0:
                temp=str(self.ArduinoSerial.readline())

                #print(temp)
                temp=temp[2:4]
                flag=flag+1
                temp=""
                #time.sleep(1)
                
            else:
                temp=str(self.ArduinoSerial.readline())
                if temp[2:4]=="ID":
                    self.rfid(temp)
                    continue;
                temp=temp[2:4]
                #print(temp)
                p1s=temp[0]
                p2s=temp[1]
                #print(p1s)
                #print(p1c)
                if flag2>4:
                    if p1s!=p1c:
                        print("Parking 2 status changed!")
                        if p1s==str(1):
                            print("Parking 2 is now free")
                            if self.ps[1]==1:
                                self.ps[1]=0
                                self.sheet_obj["P2"].value=0
                                self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
                            totalfree=totalfree+1
                            if p1flag==0:
                                p1flag=p1flag+1    
                                
                        elif p1s==str(0):
                            print("Parking 2 is now occupied")
                            self.ps[1]=1
                            self.sheet_obj["P2"].value=1
                            self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
                            totalfree=totalfree-1
                        p1c=p1s
                        
                    if p2s!=p2c:
                        print("Parking 1 status changed!")
                        if p2s==str(2):
                            print("Parking 1 is now free")
                            if self.ps[0]==1:
                                self.ps[0]=0
                                self.sheet_obj["P1"].value=0
                                self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")

        
                            totalfree=totalfree+1
                            if p2flag==0:
                                p2flag=p2flag+1
                            
                                
                        elif p2s==str(9):
                            print("Parking 1 is now occupied")
                            self.ps[0]=1
                            self.sheet_obj["P1"].value=1
                            self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
                            totalfree=totalfree-1
                        p2c=p2s
            flag2=flag2+1
            time.sleep(1)
            self.rowval=0
            self.name=""

#Main
obj1=project()
obj1.mainmenu()
