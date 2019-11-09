import serial #Serial imported for Serial communication
import time #Required to use delay functions
from twilio.rest import Client
import datetime
import openpyxl


class project(object):
    def freerowcheck(self,col):
            i=1
            temp=" "
            while temp!=None:
                i=i+1
                temp=self.sheet_obj[col+str(i)].value
            return i


    def rfid(self,temp):
        #b'ID:  AB 13 06 7512\r\n'
        temp=temp[7:18]
        temp=temp.replace(" ","")
        print(temp)
        self.name=self.rowc(temp)
        
        if self.rowval!=0:
            print("USER ALREADY EXISTS! USE ANOTHER CARD!")
        #ArduinoSerial.write("USD200".encode())
        else:
            self.ArduinoSerial.write("Hello Guest".encode())
            time.sleep(1)
            self.ArduinoSerial.write("Plz Register".encode())
            self.adduser(temp)   
        time.sleep(5)

    def adduser(self,temp):
        frow=self.freerowcheck("A")
        self.sheet_obj["A"+str(frow)].value=temp
        name=str(input("Enter Name: "))
        self.sheet_obj["B"+str(frow)].value=name
        cred=int(input("Enter Initial Credit: "))
        self.sheet_obj["I"+str(frow)].value=cred
        self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
        print("Saved Successfully!")
        
    def entry(self):
        datetimeFormat = '%Y-%m-%d %H:%M:%S.%f'
        if self.sheet_obj["F"+str(self.rowval)].value==0 or self.sheet_obj["F"+str(self.rowval)].value==None:
            self.sheet_obj["D"+str(self.rowval)].value=datetime.datetime.now()
            self.sheet_obj["F"+str(self.rowval)].value=1
            name="Hello "+self.name
            print(name)
            self.ArduinoSerial.write(name.encode())
            time.sleep(1)
            credit="Credit: Rs"+str(self.sheet_obj["I"+str(self.rowval)].value)
            self.ArduinoSerial.write(credit.encode())
            
        elif self.sheet_obj["F"+str(self.rowval)].value==1:
            self.sheet_obj["E"+str(self.rowval)].value=datetime.datetime.now()
            s=datetime.datetime.now()
            diff = datetime.datetime.strptime(str(s), datetimeFormat)- datetime.datetime.strptime(str(self.sheet_obj["D"+str(self.rowval)].value), datetimeFormat)
            name="Bye "+self.name
            self.ArduinoSerial.write(name.encode())
            time.sleep(1)
            print("Difference:", int(diff.seconds)//60)
            self.sheet_obj["G"+str(self.rowval)].value=int(diff.seconds)/60
            self.sheet_obj["F"+str(self.rowval)].value=0
            
            time1="Parked: "+str(int(diff.seconds)//60)+"Mins"
            self.ArduinoSerial.write(time1.encode())
        self.wb_obj.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
            
        

    def rowc(self,uid):
        i=1
        temp=" "
        while temp!=None:
            i=i+1
            temp=self.sheet_obj["A"+str(i)].value
            if temp==uid:
                self.rowval=i
                return self.sheet_obj["B"+str(i)].value
            else:
                return "Guest"
        
        
    def __init__ (self):
        self.mode=0
        self.acc=" "
        self.auth=[]
        self.username=" "
        self.otp=0
        self.cost=0
        self.wb=" "
        self.sheet_obj=" "
        self.wb_obj=" "
        self.location=" "
        self.access=" "
        self.rowval=0
    
#Main

    def mainmenu(self):
        ArduinoSerial = serial.Serial('com11',9600) #Create Serial port object called arduinoSerialData
        self.ArduinoSerial = ArduinoSerial
        time.sleep(2) #wait for 2 secounds for the communication to get established

        try:
            path="C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx"
            self.wb_obj = openpyxl.load_workbook(path)
            self.sheet_obj = self.wb_obj.active
            
        except:
            wb = openpyxl.Workbook()
            sheet=wb.active
            sheet['A1'].value="UID"
            sheet['B1'].value="Name"
            sheet['D1'].value="Entry"
            sheet['E1'].value="Exit"
            sheet['G1'].value="Diff"
            sheet['F1'].value="Status"
            sheet['I1'].value="Credit"
            wb.save("C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx")
            path="C:\\Users\\USERNAME\\Desktop\\"+"txn.xlsx"
            self.wb_obj = openpyxl.load_workbook(path)
            self.sheet_obj = self.wb_obj.active
            

        #print ("Test program - Arduino to Python connection.")
        p1c=100
        p2c=100
        flag=-1
        p1flag=0
        p2flag=0
        totalfree=0
        print("--- Scan a card to add a new user ----")
        while 1: #Do this forever
            
            status=str(totalfree)
            #ArduinoSerial.write(status.encode())
            if flag<0:
                temp=str(ArduinoSerial.readline())
                
                #print(temp)
                temp=temp[2:4]
                flag=flag+1
                temp=""
                #time.sleep(1)
            if flag==2:
                print("--- Scan a card to add a new user ----")
                flag=0
                
            else:
                
                temp=str(ArduinoSerial.readline())
                if temp[2:4]=="ID":
                    
                    self.rfid(temp)
                    flag=2;
                    continue;
            
            time.sleep(1)
            self.rowval=0

#Main
obj1=project()
obj1.mainmenu()
